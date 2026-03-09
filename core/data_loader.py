"""数据表加载器 - 支持动态函数 + JSON"""
import openpyxl
import json
from datetime import datetime
import random
import string

class DataLoader:
    FUNCTIONS = {
        'now': lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'today': lambda: datetime.now().strftime('%Y-%m-%d'),
        'timestamp': lambda: int(datetime.now().timestamp()),
        'random_str': lambda n=8: ''.join(random.choices(string.ascii_letters, k=int(n))),
        'random_int': lambda min=1, max=100: random.randint(int(min), int(max)),
        'random_email': lambda: f"test_{random.randint(1000,9999)}@example.com",
    }
    
    def load(self, data_path: str):
        if '::' in data_path:
            parts = data_path.split('::')
            file = parts[0]
            if file.endswith('.xlsx'):
                return self.load_excel(file, parts[1], parts[2])
            elif file.endswith('.json'):
                return self.load_json(file, parts[1])
        return {}
    
    def load_excel(self, file_path: str, sheet: str, row_id: str):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == row_id:
                data = dict(zip(headers, [cell.value for cell in row]))
                return self._process_functions(data)
        return {}
    
    def load_json(self, file_path: str, data_id: str):
        with open(file_path, 'r', encoding='utf-8') as f:
            data_list = json.load(f)
        for item in data_list:
            if item.get('DataID') == data_id:
                return self._process_functions(item)
        return {}
    
    def _process_functions(self, data: dict) -> dict:
        result = {}
        for key, value in data.items():
            if isinstance(value, str) and value.startswith('${') and value.endswith('}'):
                func_expr = value[2:-1]
                result[key] = self._eval_function(func_expr)
            else:
                result[key] = value
        return result
    
    def _eval_function(self, expr: str):
        if '(' in expr:
            func_name = expr[:expr.index('(')]
            args_str = expr[expr.index('(')+1:expr.index(')')]
            args = [a.strip() for a in args_str.split(',')] if args_str else []
            return self.FUNCTIONS[func_name](*args)
        return self.FUNCTIONS[expr]()
